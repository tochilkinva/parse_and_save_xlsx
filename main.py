import os.path

from openpyxl import Workbook, load_workbook

from parse import get_page, get_vacancies, parse_vacancies

FILE_NAME = 'vacancies.xlsx'
URL = '''https://hh.ru/search/vacancy?area=1&fromSearchLine=
true&st=searchVacancy&text=Python+junior&from=suggest_post'''


def init(file_name=FILE_NAME):
    """Создаем xlsx файл с нужными заголовками"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'Дата', 'Название', 'Зарплата',  'Работодатель',
                      'Обязанности', 'Требования', 'Ссылка'])
    workbook.save(file_name)
    workbook.close()


def text_save(text):
    """Сохраняем текст в файл"""
    file = open(r"text.txt", "w", encoding="utf-8")
    file.write(str(text))
    file.close()


def text_load():
    """Загружаем текст из файла и возвращаем"""
    file = open(r"text.txt", "r", encoding="utf-8")
    text = file.read()
    file.close()
    return eval(text)


def get_vacancies_one_page(url):
    """Получаем вакансии из одной страницы и сохраняем в файл"""
    if not os.path.exists(FILE_NAME):
        init(FILE_NAME)

    workbook = load_workbook(filename=FILE_NAME)
    worksheet = workbook.active

    html = get_page(url)
    vacancies = parse_vacancies(html)

    for key in vacancies:
        worksheet.append([key,
                          vacancies[key]['Дата'],
                          vacancies[key]['Название'],
                          vacancies[key]['Зарплата'],
                          vacancies[key]['Работодатель'],
                          vacancies[key]['Обязанности'],
                          vacancies[key]['Требования'],
                          vacancies[key]['Cсылка']
                          ])

    workbook.save(FILE_NAME)
    workbook.close()


def get_vacancies_many_pages(url):
    """Получаем вакансии из нескольких страниц и сохраняем в файл"""
    if not os.path.exists(FILE_NAME):
        init(FILE_NAME)

    workbook = load_workbook(filename=FILE_NAME)
    worksheet = workbook.active

    vacancies = get_vacancies(url)

    for key in vacancies:
        worksheet.append([key,
                          vacancies[key]['Дата'],
                          vacancies[key]['Название'],
                          vacancies[key]['Зарплата'],
                          vacancies[key]['Работодатель'],
                          vacancies[key]['Обязанности'],
                          vacancies[key]['Требования'],
                          vacancies[key]['Cсылка']
                          ])

    workbook.save(FILE_NAME)
    workbook.close()


if __name__ == "__main__":
    get_vacancies_one_page(URL)
    # get_vacancies_many_pages(URL)
    print('Список вакансий создан!')
